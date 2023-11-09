from django.shortcuts import render, redirect, get_object_or_404
from principal.models import Curso
from django.contrib.auth.decorators import user_passes_test, login_required
from django.contrib.auth.models import User  # O el modelo de usuario que estés utilizando
from django.contrib import messages
from django.contrib.auth.decorators import user_passes_test
from principal.models import Nota, Curso
from .forms import EditarNotaForm
from django.http import HttpResponse
from openpyxl import Workbook
from django.views.generic import View

###########################################################################
def editar_nota(request, nota_id):
    nota = get_object_or_404(Nota, id=nota_id)
    if request.method == 'POST':
        form = EditarNotaForm(request.POST, instance=nota)
        if form.is_valid():
            form.save()
            # Redirige a la página de detalles del estudiante después de editar la nota
            return redirect('ver_detalles_estudiante', user_id=nota.estudiante.id)
    else:
        form = EditarNotaForm(instance=nota)
    return render(request, 'docente/editar_curso.html', {'form': form, 'nota': nota})
###########################################################################
def exportar_a_excel(request):
    docente = request.user  # Obtener el usuario docente actual
    cursos_asignados = Curso.objects.filter(docente=docente)

    # Crear un nuevo libro de Excel
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Cursos y Estudiantes"

    # Definir las cabeceras
    worksheet.cell(row=1, column=1, value="Curso")
    worksheet.cell(row=1, column=2, value="Estudiante")
    worksheet.cell(row=1, column=3, value="Nota")

    row = 2

    # Llenar el libro de Excel con los datos
    for curso in cursos_asignados:
        estudiantes_inscritos = curso.estudiantes_inscritos.all()
        for estudiante in estudiantes_inscritos:
            nota = Nota.objects.filter(estudiante=estudiante).first()  # Obtener la primera nota (ajusta según tus necesidades)
            if nota:
                worksheet.cell(row=row, column=1, value=curso.nombre)
                worksheet.cell(row=row, column=2, value=f"{estudiante.first_name} {estudiante.last_name}")
                worksheet.cell(row=row, column=3, value=nota.valor)  # Ajusta según tu modelo de Nota
                row += 1

    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = "attachment; filename=cursos_y_estudiantes.xlsx"
    workbook.save(response)

    return response
###########################################################################






















@login_required
def cursos_asignados(request):
    docente = request.user  # Obtener el usuario docente actual
    # Filtrar los cursos que han sido asignados a este docente
    cursos_asignados = Curso.objects.filter(docente=docente)
    context = {
        'cursos_asignados': cursos_asignados,
    }
    return render(request, 'docente/cursos_asignados.html', context)

def estudiantes_inscritos(request):
    docente = request.user  # Obtener el usuario docente actual
    # Filtrar los cursos asignados a este docente
    cursos_asignados = Curso.objects.filter(docente=docente)

    cursos_con_estudiantes = []

    # Obtener un conjunto de estudiantes inscritos para cada curso
    for curso in cursos_asignados:
        estudiantes_inscritos = curso.estudiantes_inscritos.all()
        cursos_con_estudiantes.append({
            'curso': curso,
            'estudiantes_inscritos': estudiantes_inscritos
        })

    context = {
        'cursos_con_estudiantes': cursos_con_estudiantes,
    }

    return render(request, 'docente/estudiantes_inscritos.html', context)

#    docente = request.user  # Obtener el usuario docente actual
#    cursos_asignados = Curso.objects.filter(docente=docente)     # Filtrar los cursos asignados a este docente
#    estudiantes_inscritos = []     # Crear una lista para almacenar a los estudiantes inscritos en los cursos asignados
#    for curso in cursos_asignados:    # Iterar a través de los cursos y obtener a los estudiantes inscritos
#        estudiantes = curso.estudiantes_inscritos.all()
#        estudiantes_inscritos.extend(estudiantes)
#        context = {
#        'cursos_asignados': cursos_asignados,
#        'estudiantes_inscritos': estudiantes_inscritos,
#    }

#    return render(request, 'docente/estudiantes_inscritos.html', context)

def docente_puede_ver_detalles(user):
    return user.groups.filter(name='Docente').exists()

@user_passes_test(docente_puede_ver_detalles)
def ver_detalles_estudiante(request, user_id):
    estudiante = get_object_or_404(User, id=user_id)
    nota = Nota.objects.filter(estudiante=estudiante).first()
    return render(request, 'docente/detalles_estudiante.html', {'estudiante': estudiante, 'nota':nota})

def cuenta2(request):
    cursos = Curso.objects.all()
    user_belongs_to_docente_group = request.user.groups.filter(name='Docente').exists()
    context = {
        'user_belongs_to_docente_group': user_belongs_to_docente_group,
        'cursos': cursos,
    }
    return render(request, 'docente/cuenta.html', context)

def perfil(request):#ver el perfil del usuario 
    usuario = request.user
    grupos_del_usuario = usuario.groups.all()
    contexto = {
        'usuario': usuario,
        'grupos_del_usuario': grupos_del_usuario,
    }
    return render(request, 'docente/perfil.html', contexto)


def lista_de_cursos2(request):
    cursos = Curso.objects.all()
    for curso in cursos:
        # Calcula la cantidad de cupos disponibles restando la cantidad de estudiantes inscritos al cupo total
        curso.cupos_disponibles = curso.cupo - curso.cantidad_estudiantes
    data = {
        'cursos': cursos
    }
    return render(request, 'docente/lista_cursos.html', data)
