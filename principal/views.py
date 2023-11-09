from django.shortcuts import render, redirect, get_object_or_404
from .models import Curso #desde models.py se importa la clase Curso
from .forms import ContactoForm, CursoForm, CustormUserCreationForm
from django.contrib.auth import authenticate, login 
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.models import Group, User
from django.http import HttpResponse
from openpyxl import Workbook
from django.views.generic import View
from .models import Curso  # Asegúrate de importar tu modelo
from django.shortcuts import render, redirect
from django.contrib.auth import login, authenticate
from django.contrib.auth.forms import  AuthenticationForm
from django.urls import reverse
# Create your views here.

###########################################################################
class ExportToExcelView(View):# para descargar archivo excel
   def get(self, request):
       workbook = Workbook()
       sheet = workbook.active

       # Definir los encabezados
       headers = ['Nombre', 'Precio', 'Descripcion', 'Precio en Q', 'Cupo', 'Estudiantes inscritos', 'Docente']
       for col_num, header in enumerate(headers, 1):
           cell = sheet.cell(row=1, column=col_num)
           cell.value = header

       cursos = Curso.objects.all()
       row_num = 2
       for curso in cursos:
           # Llena las celdas de datos de acuerdo a tus necesidades
           sheet.cell(row=row_num, column=1, value=curso.nombre)
           sheet.cell(row=row_num, column=2, value=curso.precio)
           sheet.cell(row=row_num, column=3, value=curso.descripcion)
           sheet.cell(row=row_num, column=4, value=curso.precio)
           sheet.cell(row=row_num, column=5, value=curso.cupo)
           sheet.cell(row=row_num, column=6, value=curso.cantidad_estudiantes)
           
           docente = curso.docente
           if docente:
            docente_nombre = f"{docente.first_name} {docente.last_name}"
            sheet.cell(row=row_num, column=7, value=docente_nombre)
            
            # Agrega más celdas para otras columnas si es necesario


            row_num += 1

       response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
       response['Content-Disposition'] = 'attachment; filename=listado_de_cursos.xlsx'

       workbook.save(response)
       return response
###########################################################################

###########################################################################
def asignar_docente_a_curso(request, curso_id):
    if request.method == 'POST':
        curso = get_object_or_404(Curso, id=curso_id)
        docente_id = request.POST.get('docente_id')
        if docente_id:
            docente = get_object_or_404(User, id=docente_id)
            curso.docente = docente
            curso.save()
    return redirect('detalle_curso', curso_id=curso_id)
###########################################################################

###########################################################################
def inicio(request):#funcion con nombre, 
    return render(request, 'principal/inicio.html')#Para encontrar el html
###########################################################################

###########################################################################
def Escuelas(request):
    return render(request, "principal/escuelas.html")
###########################################################################

###########################################################################
def contacto(request):
    data = {
        'form':ContactoForm()
    }
    if request.method == 'POST':
        formulario = ContactoForm(data = request.POST)
        if formulario.is_valid(): #si el formulario es valido
            formulario.save() #se guarda el formulario
            data["mensaje"] = "Contacto enviado" #muestra mensaje
        else:
            data["form"] = formulario
    return render(request, 'principal/contacto.html', data)
###########################################################################

###########################################################################
def cursos(request):
    curso = Curso.objects.all()
    data = {
        'cursos': curso
    }
    return render(request, 'principal/cursos.html', data)
###########################################################################

@permission_required('principal.add_curso')#para que solo aparezca a los que tengan los permisos
def agregar_curso(request):
    data = {
        'form':CursoForm
    }
    if request.method == 'POST':
        formulario = CursoForm(data=request.POST, files= request.FILES)
        if formulario.is_valid():
            formulario.save()
            data["mensaje"] = "Guardado correctamente"
        else:
            data["form"] = formulario
    return render(request, 'agregarcurso/agregar.html', data)

@permission_required('principal.view_curso')
def listar_cursos(request):
    cursos = Curso.objects.all()
    data = {
        'cursos': cursos
    }
    return render(request, 'agregarcurso/listar.html', data)

@permission_required('principal.change_curso')
def modificar_curso(request, id):
    curso = get_object_or_404(Curso, id=id)
    data={
        'form':CursoForm(instance=curso)
    }
    if request.method == 'POST':
        formulario = CursoForm(data=request.POST, instance = curso, files=request.FILES)
        if formulario.is_valid():
            formulario.save()
            messages.success(request,"Modificado correctamente")
            return redirect(to="listar_cursos")
        data["form"]=formulario
    return render(request, 'agregarcurso/modificar.html',data)


@permission_required('principal.delete_curso')
def eliminar_curso(request, id):
    curso = get_object_or_404(Curso, id=id)
    curso.delete()
    return redirect(to= "listar_cursos")

def registro(request):
    data = {
        'form': CustormUserCreationForm()
    }
    if request.method == 'POST':
        formulario = CustormUserCreationForm(data=request.POST)
        if formulario.is_valid():
            formulario.save()
            user = authenticate(username= formulario.cleaned_data["username"], password=formulario.cleaned_data["password1"])

            grupo_estudiante, creado = Group.objects.get_or_create(name='Estudiante')
            user.groups.add(grupo_estudiante)

            login(request, user)
            messages.success(request, "Te has registrado correctamente")
            return redirect('cuenta')
        data["form"]= formulario
    return render(request, 'registration/registro.html', data)

@login_required
def registro2(request):
    data = {
        'form': CustormUserCreationForm()
    }
    if request.method == 'POST':
        formulario = CustormUserCreationForm(data=request.POST)
        if formulario.is_valid():
            formulario.save()
            user = authenticate(username= formulario.cleaned_data["username"], password=formulario.cleaned_data["password1"])

            grupo_docente, creado = Group.objects.get_or_create(name='Docente')
            user.groups.add(grupo_docente)

            messages.success(request, "Te has registrado correctamente")
            return redirect('cuenta')
        data["form"]= formulario
    contexto = {
        'es_staff': request.user.is_staff, 
    }

    data.update(contexto)

    return render(request, 'registration/registro.html', data)

def custom_login(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)

        if user is not None:
            if user.is_staff:
                # El usuario es un administrador, inicia sesión y redirige al panel de administración
                login(request, user)
                return redirect('admin:index')  # Redirige al panel de administración

    return render(request, 'registration/administrativo.html')